#!/usr/bin/env python3
"""Inventaría encabezados físicos de las tres fuentes Excel sin modificarlas.

El resultado es una evidencia reproducible de cobertura. Las reglas clasifican
terminología operativa; nunca importan filas ni convierten totales en captura.
"""

from __future__ import annotations

import argparse
import csv
import re
import unicodedata
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


SPECS = {
    "PROYECTOS VÍAS SEGUIMIENTO GENERAL.xlsx": {
        "INFORME GENERAL": (2, 4, 1, None),
        "COP´S COLECTIVOS": (2, 4, 1, None),
        "INDIVIDUALES": (2, 3, 1, None),
    },
    "SEGUIMIENTO DE ACTIVIDADES LIBERACIÓN DE VIAS-INDIVIDUALES-MQ.xlsx": {
        "General ": (7, 8, 1, None),
        "PROPUESTA": (3, 5, 1, None),
        "COP´S PENDIENTES": (3, 5, 1, None),
        "Hoja2": (1, 1, 1, 0),
        "Hoja1": (1, 2, 2, 12),
    },
    "Copia de SEGUIMIENTO DE ACTIVIDADES LIBERACIÓN DE VIAS (REV) MQ.xlsx": {
        "General ": (7, 8, 1, None),
        "INFORME M-Q": (4, 6, 1, None),
        "ORV": (3, 5, 1, None),
        "ASAMBLEAS PENDIENTES": (3, 5, 1, None),
        "PCOLECTIVAS": (1, 2, 1, 11),
    },
}

KPI_SHEETS = {"INFORME GENERAL", "COP´S COLECTIVOS", "INDIVIDUALES", "PCOLECTIVAS", "Hoja1"}


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\n", " ")).strip()


def key(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return re.sub(r"[^A-Z0-9]+", " ", normalized.encode("ascii", "ignore").decode().upper()).strip()


def merged_value(sheet, row: int, column: int) -> str:
    value = sheet.cell(row, column).value
    if value not in (None, ""):
        return clean(value)
    for area in sheet.merged_cells.ranges:
        if area.min_row <= row <= area.max_row and area.min_col <= column <= area.max_col:
            return clean(sheet.cell(area.min_row, area.min_col).value)
    return ""


def last_header_column(sheet, first_row: int, last_row: int) -> int:
    columns = [
        column
        for column in range(1, sheet.max_column + 1)
        if any(sheet.cell(row, column).value not in (None, "") for row in range(first_row, last_row + 1))
    ]
    return max(columns, default=0)


def coverage(path: str, sheet: str, exact: str):
    text = key(path)
    exact_key = key(exact)
    individual = "individual" if "INDIVIDUAL" in key(sheet) or "PARCELA" in text or "TITULAR" in text else "general"

    if exact_key in {"NUM", "NO", "NUMERO CONSECUTIVO"} and "SOLICITUD" not in text:
        return ("orden de hoja", "—", "—", "—", "No se captura", "NO_APLICA", individual)
    if "TRIMESTRE" in exact_key or "MES A CONTAR" in exact_key:
        return ("periodo derivado de fecha", "vista KPI", "DashboardKPI", "/api/dashboard", "Dashboard", "DERIVADO", individual)
    if "TRAMO" in text:
        return ("referencia histórica del tramo", "proyecto_nucleo_referencia.tipo/valor", "ProyectoNucleoReferencia", "/api/proyecto-nucleo/{id}/referencias", "General · Referencias del tramo", "REFERENCIA_LEGACY", "general")
    if sheet in KPI_SHEETS and "TOTAL DE NUCLEOS" in text:
        return ("total de núcleos derivado", "vw_dashboard_kpi", "DashboardKPI", "/api/dashboard", "Dashboard", "DERIVADO", "general")
    if sheet in KPI_SHEETS and not any(token in text for token in ("ENTIDAD", "MUNICIPIO", "NUCLEO", "PROYECTO", "RESIDENCIA")):
        return ("indicador agregado", "vw_dashboard_kpi", "DashboardKPI", "/api/dashboard", "Dashboard", "DERIVADO", individual)
    if "ACUSE FIFONAFE" in text:
        return ("fecha de acuse FIFONAFE", "tramite_fifonafe.acuse_fifonafe_fecha", "TramiteFifonafe", "/api/proyecto-nucleo/{id}/fifonafe", "FIFONAFE · Acuse", "OK", individual)

    rules = [
        (("ENTIDAD",), "entidad federativa", "entidad_federativa.nombre", "EntidadFederativa", "/api/catalogos/entidades", "Selector de contexto · ENTIDAD", "OK", "general"),
        (("MUNICIPIO",), "municipio", "municipio.nombre", "Municipio", "/api/catalogos/municipios", "Selector de contexto · MUNICIPIO", "OK", "general"),
        (("RESIDENCIA",), "residencia administrativa", "proyecto_nucleo.residencia", "ProyectoNucleo", "/api/proyecto-nucleo/{id}", "General · RESIDENCIA", "OK", "general"),
        (("CONSECUTIVO",), "consecutivo operativo", "proyecto_nucleo_referencia.valor", "ProyectoNucleoReferencia", "/api/proyecto-nucleo/{id}/referencias", "General · CONSECUTIVO", "OK", "general"),
        (("NUCLEO AGRARIO",), "núcleo agrario", "nucleo_agrario.nombre", "NucleoAgrario", "/api/proyectos/{id}/nucleos", "Selector de contexto · NÚCLEO AGRARIO", "OK", "general"),
        (("E C",), "ejido o comunidad", "nucleo_agrario.tipo_nucleo", "NucleoAgrario", "/api/proyectos/{id}/nucleos", "General · E/C", "OK", "general"),
        (("ORGANIZADORA",), "persona organizadora agraria responsable", "proyecto_nucleo.responsable_nombre", "ProyectoNucleo", "/api/proyecto-nucleo/{id}", "General · PERSONA ORGANIZADORA RESPONSABLE", "OK", "general"),
        (("CONTACTO",), "teléfono de contacto", "proyecto_nucleo.responsable_telefono", "ProyectoNucleo", "/api/proyecto-nucleo/{id}", "General · DATOS DE CONTACTO (TELÉFONO)", "OK", "general"),
        (("PROYECTO",), "proyecto ferroviario", "proyecto.nombre_proyecto", "Proyecto", "/api/proyectos", "Selector de contexto · Proyecto", "OK", "general"),
        (("DESTINO",), "destino de la superficie", "afectacion.destino_superficie", "Afectacion", "/api/proyecto-nucleo/{id}/afectaciones", "Afectación · DESTINO DE LA SUPERFICIE", "OK", individual),
        (("COMISARIADO",), "integrante nominal del comisariado", "orv_integrante.cargo + persona", "ORVIntegrante", "/api/orv/{id}/integrantes", "ORV y Padrón · Comisariado", "OK", "general"),
        (("CONSEJO", "VIGILANCIA"), "integrante nominal del consejo de vigilancia", "orv_integrante.cargo + persona", "ORVIntegrante", "/api/orv/{id}/integrantes", "ORV y Padrón · Consejo de Vigilancia", "OK", "general"),
        (("ORV",), "vigencia o estado de ORV", "orv.inicio_vigencia/fin_vigencia/acta_inscrita_ran", "ORV", "/api/nucleos/{id}/orv", "ORV y Padrón", "OK", "general"),
        (("PADRON",), "padrón vigente", "padron_historial.fecha_padron/numero_sujetos", "PadronHistorial", "/api/nucleos/{id}/padrones", "ORV y Padrón · Padrón", "OK", "general"),
        (("PROGRAMADA POR NA",), "marca operativa por núcleo agrario", "derivado/documentado desde actividad", "ActividadCampo", "/api/proyecto-nucleo/{id}/actividades", "Actividades · contexto operativo", "DERIVADO", individual),
        (("REALIZADA POR NA",), "marca operativa por núcleo agrario", "derivado/documentado desde actividad", "ActividadCampo", "/api/proyecto-nucleo/{id}/actividades", "Actividades · contexto operativo", "DERIVADO", individual),
        (("SENSIBILIZ",), "actividad de sensibilización", "actividad_campo", "ActividadCampo", "/api/proyecto-nucleo/{id}/actividades", "Actividades · Sensibilización", "OK", "general"),
        (("CAMINAMIENTO",), "actividad de caminamiento", "actividad_campo", "ActividadCampo", "/api/proyecto-nucleo/{id}/actividades", "Actividades · Caminamiento", "OK", individual),
        (("PARCELA",), "identificación de parcela", "parcela.numero_parcela/tipo_parcela", "Parcela", "/api/nucleos/{id}/parcelas", "Afectación individual · Parcela", "OK", "individual"),
        (("TITULAR",), "titular de parcela", "parcela_titular + persona", "ParcelaTitular", "/api/parcelas/{id}/titulares", "Afectación individual · Titular", "OK", "individual"),
        (("CONSTANCIA", "VIGENCIA"), "constancia de vigencia de derechos", "parcela.constancia_vigencia_fecha + documento", "Parcela", "/api/nucleos/{id}/parcelas", "Parcelas · Vigencia de derechos", "OK", "individual"),
        (("CERTIFICADO", "PARCELARIO"), "certificado parcelario", "parcela.certificado_parcelario", "Parcela", "/api/nucleos/{id}/parcelas", "Parcelas · Certificado parcelario", "OK", "individual"),
        (("FOLIO", "DERECHOS"), "folio de derechos", "parcela.folio_derechos", "Parcela", "/api/nucleos/{id}/parcelas", "Parcelas · Folio de derechos", "OK", "individual"),
        (("SUPERFICIE",), "superficie administrativa en hectáreas", "afectacion/convenio.superficie_*_ha", "Afectacion/Convenio", "/api/proyecto-nucleo/{id}/afectaciones", "Afectación o Convenio · Superficie (ha)", "OK", individual),
        (("AVALUO",), "avalúo administrativo", "afectacion.avaluo_monto/fecha/referencia/institucion", "Afectacion", "/api/proyecto-nucleo/{id}/afectaciones", "Afectación · Avalúo", "OK", "general"),
        (("EXPROPIACION DIRECTA",), "condición de expropiación directa", "afectacion.condicion_especial", "Afectacion", "/api/proyecto-nucleo/{id}/afectaciones", "Afectación · Condición especial", "OK", "colectivo"),
        (("NO AFECTA TIERRAS DE USO COMUN",), "proyecto sin afectación a uso común", "afectacion.condicion_especial", "Afectacion", "/api/proyecto-nucleo/{id}/afectaciones", "Afectación · Condición especial", "OK", "colectivo"),
        (("COMUNIDAD INDIGENA",), "condición de comunidad indígena", "afectacion.condicion_especial/descripcion_condicion", "Afectacion", "/api/proyecto-nucleo/{id}/afectaciones", "Afectación · Condición especial", "OK", "colectivo"),
        (("ASAMBLEA",), "seguimiento de Asamblea", "asamblea", "Asamblea", "/api/proyecto-nucleo/{id}/asambleas", "Asambleas", "OK", "colectivo"),
        (("CONVOCATORIA",), "convocatoria de Asamblea", "asamblea.primera_convocatoria/segunda_convocatoria", "Asamblea", "/api/proyecto-nucleo/{id}/asambleas", "Asambleas · Convocatorias", "OK", "colectivo"),
        (("MONTO 90",), "monto 90 por ciento", "convenio.monto_90", "Convenio", "/api/convenios/{id}", "Convenios · MONTO 90%", "OK", individual),
        (("MONTO 100",), "monto 100 por ciento", "convenio.monto_100", "Convenio", "/api/convenios/{id}", "Convenios · MONTO 100%", "OK", individual),
        (("MONTO BDT",), "monto BDT", "convenio.monto_bdt", "Convenio", "/api/convenios/{id}", "Convenios · MONTO BDT", "OK", individual),
        (("CONVENIO",), "seguimiento de convenio", "convenio", "Convenio", "/api/afectaciones/{id}/convenios", "Convenios", "OK", individual),
        (("RAN",), "seguimiento registral RAN", "asamblea/convenio fechas, solicitud y calificación", "Asamblea/Convenio", "/api/proyecto-nucleo/{id}/asambleas", "Asambleas o Convenios · Seguimiento RAN", "OK", individual),
        (("SOLICITUD",), "número de solicitud", "asamblea/convenio.numero_solicitud_ingreso", "Asamblea/Convenio", "/api/proyecto-nucleo/{id}/asambleas", "Seguimiento RAN · NÚMERO DE SOLICITUD", "OK", individual),
        (("CALIFICACION",), "calificación registral", "asamblea/convenio.calificacion_registral", "Asamblea/Convenio", "/api/proyecto-nucleo/{id}/asambleas", "Seguimiento RAN · CALIFICACIÓN REGISTRAL", "OK", individual),
        (("FIFONAFE",), "informe de no conflictos FIFONAFE", "tramite_fifonafe", "TramiteFifonafe", "/api/proyecto-nucleo/{id}/fifonafe", "FIFONAFE", "OK", individual),
        (("OFICIO",), "oficio operativo", "tramite_fifonafe oficios o documento", "TramiteFifonafe/Documento", "/api/proyecto-nucleo/{id}/fifonafe", "FIFONAFE · Secuencia de oficios", "OK", individual),
        (("CONFLICT",), "resultado de conflictos", "tramite_fifonafe.hay_conflictos/resultado_no_conflictos", "TramiteFifonafe", "/api/proyecto-nucleo/{id}/fifonafe", "FIFONAFE · Conflictos", "OK", individual),
        (("INDEMNIZ",), "indemnización", "indemnizacion", "Indemnizacion", "/api/afectaciones/{id}/indemnizacion", "Indemnización y Pagos", "OK", individual),
        (("PAGO",), "pago de indemnización", "pago", "Pago", "/api/indemnizaciones/{id}/pagos", "Indemnización y Pagos", "OK", individual),
        (("SOPORTE",), "soporte documental", "documento/documento_version/documento_vinculo", "Documento", "/api/documentos", "Documentos", "OK", individual),
        (("OBSERV",), "observaciones operativas", "observaciones de la entidad correspondiente", "Schema de entidad", "endpoint del proceso", "Detalle del proceso · Observaciones", "OK", individual),
    ]
    for tokens, meaning, table, schema, endpoint, screen, status, scope in rules:
        if all(token in text for token in tokens):
            return (meaning, table, schema, endpoint, screen, status, scope)

    if "VALIDACION PA SICT" in text:
        return ("validación operativa PA/SICT", "sin atributo inequívoco", "—", "—", "FIFONAFE", "DECISION_FUNCIONAL_REQUERIDA", individual)
    if "ENTREGA" in text and "EXPEDIENTE" in text:
        return ("entrega de expediente SICT–PA", "indemnizacion.fecha_entrega_expediente_pa", "Indemnizacion", "/api/afectaciones/{id}/indemnizacion", "Indemnización y Pagos", "OK", individual)
    if "FECHA" in text or "ESTATUS" in text or "RESULTADO" in text:
        return ("dato temporal o de estado del proceso", "entidad del bloque de encabezado", "Schema de entidad", "endpoint del proceso", "Pestaña del proceso", "OK", individual)
    return ("campo operativo por clasificar institucionalmente", "—", "—", "—", "—", "DECISION_FUNCIONAL_REQUERIDA", individual)


def audit(source: Path):
    rows = []
    for filename, sheets in SPECS.items():
        workbook = openpyxl.load_workbook(source / filename, read_only=False, data_only=False)
        for sheet_name, (first_row, last_row, first_column, last_override) in sheets.items():
            sheet = workbook[sheet_name]
            last_column = last_override if last_override is not None else last_header_column(sheet, first_row, last_row)
            for column in range(first_column, last_column + 1):
                hierarchy = []
                for row in range(first_row, last_row + 1):
                    value = merged_value(sheet, row, column)
                    if value and (not hierarchy or hierarchy[-1] != value):
                        hierarchy.append(value)
                exact = hierarchy[-1] if hierarchy else "(encabezado vacío dentro del bloque físico)"
                block = " → ".join(hierarchy[:-1]) or "General"
                path = " → ".join(hierarchy) or exact
                meaning, table, schema, endpoint, screen, status, scope = coverage(path, sheet_name, exact)
                rows.append({
                    "archivo": filename,
                    "hoja": sheet_name.strip(),
                    "columna_excel": get_column_letter(column),
                    "bloque_encabezado": block,
                    "nombre_exacto_excel": exact,
                    "significado_normalizado": meaning,
                    "ambito": scope,
                    "tabla_columna_bd": table,
                    "schema_pydantic": schema,
                    "endpoint": endpoint,
                    "pantalla_control_ui": screen,
                    "estado": status,
                })
    return rows


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--unique", action="store_true")
    args = parser.parse_args()
    rows = audit(args.source)
    if args.unique:
        for value in sorted({f"{row['bloque_encabezado']} → {row['nombre_exacto_excel']}" for row in rows}):
            print(value)
        return
    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", newline="", encoding="utf-8-sig") as target:
        writer = csv.DictWriter(target, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    print(f"{len(rows)} columnas auditadas; salida: {args.output}")


if __name__ == "__main__":
    main()
